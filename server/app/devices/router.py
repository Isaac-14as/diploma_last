from fastapi import APIRouter, Depends
from fastapi.responses import FileResponse
from openpyxl import Workbook

from app.devices.schemas import *
from app.users.dependencies import get_current_user, is_admin_user
from piccolo_db.tables import AccidentLog, Device, ManagementLog, ValueDevice
from datetime import datetime, timedelta

router = APIRouter(
    prefix="/device",
    tags=["Devices & Устройства"],
)

async def get_value_device_by_date_time(device_id: int, date: str, time_from: str, time_to: str):
    date = datetime.strptime(date, "%Y-%m-%d").date()
    if time_from and time_to and date:
        time_from = datetime.strptime(time_from, "%H:%M:%S").time()
        time_to = datetime.strptime(time_to, "%H:%M:%S").time()
        datetime_from = datetime.combine(date, time_from)
        datetime_to = datetime.combine(date, time_to)
        return await ValueDevice.objects().where(
            (ValueDevice.device_id == device_id) &
            (ValueDevice.date_of_origin >= datetime_from) & 
            (ValueDevice.date_of_origin <= datetime_to)
        ).order_by(ValueDevice.date_of_origin)
    return await ValueDevice.objects().where(
            (ValueDevice.device_id == device_id) &
            (ValueDevice.date_of_origin >= date) & 
            (ValueDevice.date_of_origin < (date + timedelta(days=1)))
        ).order_by(ValueDevice.date_of_origin)



@router.get('/get_all_devices', response_model=list[SDeviceGet], dependencies=[Depends(get_current_user)])
async def get_all_devices():
    return await Device.select()


@router.get('/get_device_by_id/{device_id}', response_model=SDeviceGet, dependencies=[Depends(get_current_user)])
async def get_device_by_id(device_id: int):
    return await Device.objects().where(Device.id == device_id).first()


@router.get('/get_value_device_by_id/{device_id}', response_model=list[SValueDeviceGet], dependencies=[Depends(get_current_user)])
async def get_value_device_by_id(device_id: int, date: str = None, time_from: str = None, time_to: str = None):
    if date:
        return await get_value_device_by_date_time(device_id, date, time_from, time_to)
    return await ValueDevice.objects().where(ValueDevice.device_id == device_id).order_by(ValueDevice.date_of_origin)


@router.get('/get_last_value_device_all', response_model=list[SValueDeviceGet], dependencies=[Depends(get_current_user)])
async def get_last_value_device_all():
    device_ids = [1, 2, 3, 4]
    res = []
    for id in device_ids:
        res.append(await ValueDevice.select().where(ValueDevice.device_id==id).order_by(ValueDevice.date_of_origin, ascending=False).first())
    return res
        


@router.get('/get_full_power/{device_id}', response_model=list[SValueFullPower], dependencies=[Depends(get_current_user)])
async def get_full_power(device_id: int, date: str = None, time_from = None, time_to = None):
    if date:
        return await get_value_device_by_date_time(device_id, date, time_from, time_to)
    return await ValueDevice.select().order_by(ValueDevice.date_of_origin)


@router.patch('/device_switch/{device_id}', dependencies=[Depends(get_current_user)])
async def device_switch(device_id: int):
    device = await Device.objects().where(Device.id==device_id).first()
    await Device.update().where(Device.id==device_id).values(is_active=not device.is_active).run()
    if device.is_active:
        return {'status': 200, 'detail': 'Устройство выключено.'}
    return {'status': 200, 'detail': 'Устройство включено.'}


@router.post('/add_management_log', dependencies=[Depends(get_current_user)])
async def add_management_log(management_log_info: SManagementAdd):
    await ManagementLog(**management_log_info.model_dump()).save()
    return {'status': 200, 'detail': 'Запиcь успешно добавлена.'}


@router.get('/get_management_log', dependencies=[Depends(get_current_user)])
async def get_management_log(date: str = None, time_from: str = None, time_to: str = None):
    user_col = [ManagementLog.user_id.email, ManagementLog.user_id.name, ManagementLog.user_id.role]
    if date and time_from and time_to:
        date = datetime.strptime(date, "%Y-%m-%d").date()
        time_from = datetime.strptime(time_from, "%H:%M:%S").time()
        time_to = datetime.strptime(time_to, "%H:%M:%S").time()
        datetime_from = datetime.combine(date, time_from)
        datetime_to = datetime.combine(date, time_to)

        return await ManagementLog.select(ManagementLog.all_columns(), user_col, ManagementLog.device_id.name).where(
            (ManagementLog.date_of_origin >= datetime_from) & 
            (ManagementLog.date_of_origin <= datetime_to)
        ).order_by(ManagementLog.date_of_origin)
    if date:
        date = datetime.strptime(date, "%Y-%m-%d").date()
        return await ManagementLog.select(ManagementLog.all_columns(), user_col, ManagementLog.device_id.name).where(
            (ManagementLog.date_of_origin >= date) & 
            (ManagementLog.date_of_origin < (date + timedelta(days=1)))
        ).order_by(ManagementLog.date_of_origin)
    return await ManagementLog.select(ManagementLog.all_columns(), user_col, ManagementLog.device_id.name).order_by(ManagementLog.date_of_origin)


@router.get('/get_accident_log/', dependencies=[Depends(get_current_user)])
async def get_accident_log(device_id: int | str = None, date: str = None, time_from: str = None, time_to: str = None):
    print(device_id)
    if device_id == 'null': 
        device_id = None
    elif device_id:
        device_id = int(device_id)
    if date == 'null':
        date = None
    if date and time_from and time_to:
        date = datetime.strptime(date, "%Y-%m-%d").date()
        time_from = datetime.strptime(time_from, "%H:%M:%S").time()
        time_to = datetime.strptime(time_to, "%H:%M:%S").time()
        datetime_from = datetime.combine(date, time_from)
        datetime_to = datetime.combine(date, time_to)
    if device_id and date  and time_from and time_to:
        return await AccidentLog.select(AccidentLog.all_columns(), AccidentLog.device_id.all_columns()).where(
            (AccidentLog.device_id == device_id) & 
            (AccidentLog.date_of_origin >= datetime_from) & 
            (AccidentLog.date_of_origin <= datetime_to)
        ).order_by(AccidentLog.date_of_origin)
    if date and time_from and time_to:
        return await AccidentLog.select(AccidentLog.all_columns(), AccidentLog.device_id.all_columns()).where(
            (AccidentLog.date_of_origin >= datetime_from) & 
            (AccidentLog.date_of_origin <= datetime_to)
        ).order_by(AccidentLog.date_of_origin)
    if device_id:
        return await AccidentLog.select(AccidentLog.all_columns(), AccidentLog.device_id.all_columns()).where(AccidentLog.device_id==device_id).order_by(AccidentLog.date_of_origin)
    return await AccidentLog.select(AccidentLog.all_columns(), AccidentLog.device_id.all_columns()).order_by(AccidentLog.date_of_origin)


@router.get("/get_reporting_device_value/{device_id}")
async def get_reporting_device_value(device_id: int, date: str = None, time_from: str = None, time_to: str = None):
    device_name = await Device.select(Device.name).where(Device.id==device_id).first()
    device_name = device_name['name']
    if date and (date != 'null'):
        res = await get_value_device_by_date_time(device_id, date, time_from, time_to)
    else:
        res = await ValueDevice.objects().where(ValueDevice.device_id==device_id)
    wb = Workbook()
    ws = wb.active
    ws.append(['Полная мощность', 'Активная мощность', 'Реактивная мощность', 'Напряжение', 'Сила тока', 'Коэффициент мощности', 'Дата/время'])
    ws.append([str(i).split(r'"')[1] for i in ValueDevice.all_columns()][1:-1])
    ws.append(['S[МВА]', 'P[МВт]', 'Q[МВАр]', 'U[кВ]', 'I[A]', 'cosφ'])
    for obj in res:
        ws.append([obj.full_power, obj.active_power, obj.reactive_power, obj.voltage, obj.amperage, obj.power_factor, obj.date_of_origin])
    filename =f'report.xlsx'
    wb.save(filename)
    return FileResponse(filename, filename=filename)


@router.get("/get_reporting_accident_log")
async def get_reporting_accident_log(device_id: int | str = None, date: str = None, time_from: str = None, time_to: str = None):
    res = await get_accident_log(device_id, date, time_from, time_to)
    wb = Workbook()
    ws = wb.active
    ws.append(['Устройство', 'Сообщение', 'Дата/время'])
    for obj in res:
        ws.append([obj['device_id.name'], obj['info'], obj['date_of_origin']])
    filename =f'report.xlsx'
    wb.save(filename)
    return FileResponse(filename, filename=filename)


@router.get("/get_reporting_management_log")
async def get_reporting_management_log(date: str = None, time_from: str = None, time_to: str = None):
    res = await get_management_log(date, time_from, time_to)
    wb = Workbook()
    ws = wb.active
    ws.append(['Устройство', 'Действие', 'Причина', 'Сотрудник', 'Дата/время'])
    for obj in res:
        ws.append([obj['device_id.name'], obj['action'], obj['info'], obj['user_id.name'] + '\n' + obj['user_id.email'] + '\n' + obj['user_id.role'], obj['date_of_origin']])
    filename =f'report.xlsx'
    wb.save(filename)
    return FileResponse(filename, filename=filename)